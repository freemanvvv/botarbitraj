"""
Модуль расчёта смет и BOQ.
Арифметика — в коде, не в LLM.
"""
import os
import re
from datetime import datetime

from .pricing_db import (
    search_materials,
    search_work,
    create_estimate,
    add_estimate_item,
    get_estimate,
    add_material,
    add_work,
    list_all_materials,
    list_all_work,
)
from .config import OUTPUT_DIR


def parse_boq_from_llm(llm_response: str) -> list[dict]:
    """
    Парсит структурированный ответ LLM в список позиций.
    Обрабатывает Markdown-таблицы с разными заголовками и диапазонами.
    """
    items = []
    lines = llm_response.split("\n")
    in_table = False

    for i, line in enumerate(lines):
        stripped = line.strip()

        # Определяем начало таблицы по строке-разделителю
        if "| ---" in stripped or "|---|---" in stripped:
            in_table = True
            continue

        if not in_table:
            continue

        # Конец таблицы
        if not stripped.startswith("|") or stripped == "|---|":
            in_table = False
            continue

        cells = [c.strip() for c in stripped.split("|") if c.strip()]

        # Пропускаем заголовок
        if any(h in cells[0].lower() for h in ["№", "n", "#", "номер", "п/п"]):
            continue
        if any(h in cells[0].lower() for h in ["наименован", "материал", "работа", "вид"]):
            continue

        # Если первая ячейка не число — пропускаем
        if not cells[0].replace(".", "").replace(",", "").isdigit():
            continue

        # Парсим: минимум 3 колонки (№, название, единица + количество)
        if len(cells) < 3:
            continue

        name = cells[1] if len(cells) > 1 else ""
        unit = cells[2] if len(cells) > 2 else ""
        qty_text = cells[3] if len(cells) > 3 else cells[2]

        quantity = _parse_quantity(qty_text)
        if quantity == 0:
            continue

        items.append({
            "name": name,
            "unit": unit,
            "quantity": quantity,
        })

    if items:
        return items

    # Fallback: парсим строки с дефисами/звёздочками
    for line in lines:
        line = line.strip()
        if not line or line.startswith("|"):
            continue

        for prefix in ["- ", "* ", "• "]:
            if line.startswith(prefix):
                content = line[len(prefix):]
                item = _parse_simple_line(content)
                if item:
                    items.append(item)
                    break

    return items


def _parse_simple_line(content: str) -> dict | None:
    """
    Парсит строку вида:
    "Фундамент ленточный — 35 м3"
    "Фундамент ленточный: 35 м3"
    """
    for sep in [" — ", " – ", " - ", ": "]:
        if sep in content:
            parts = content.split(sep, 1)
            name = parts[0].strip()
            rest = parts[1].strip()
            match = re.match(r"([\d.,]+)\s*(\S+)", rest)
            if match:
                qty = _parse_quantity(match.group(1))
                unit = match.group(2)
                if qty > 0 and unit:
                    return {"name": name, "unit": unit, "quantity": qty}
    return None


def _parse_quantity(text: str) -> float:
    """
    Парсит количество.
    Поддерживает диапазоны (берёт среднее), запятые, точки.
    """
    text = text.strip().replace(",", ".")

    # Диапазон: "35–40", "35-40", "35 - 40"
    range_match = re.match(r"([\d.]+)\s*[–\-—]\s*([\d.]+)", text)
    if range_match:
        low = float(range_match.group(1))
        high = float(range_match.group(2))
        return round((low + high) / 2, 2)

    # "~35", "≈35", "около 35"
    text = re.sub(r"^[~≈]|около\s*", "", text).strip()

    try:
        return float(text)
    except ValueError:
        return 0.0


def calculate_estimate(project_name: str, items: list[dict]) -> dict:
    """
    Рассчитывает смету.
    - Позиции материалов сопоставляются с базой расценок
    - Арифметика считается кодом
    - Результат сохраняется в SQLite
    """
    estimate_id = create_estimate(project_name)
    results = []

    for item in items:
        name = item["name"]
        quantity = item["quantity"]
        unit = item.get("unit", "шт")

        # Поиск в базе материалов
        materials = search_materials(name, limit=3)
        if materials:
            unit_price = materials[0]["price"]
            add_estimate_item(
                estimate_id, "material", name, unit, quantity, unit_price
            )
            results.append({
                "type": "material",
                "name": name,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "total": round(quantity * unit_price, 2),
            })
            continue

        # Поиск в базе работ
        work = search_work(name, limit=3)
        if work:
            unit_price = work[0]["price"]
            add_estimate_item(
                estimate_id, "work", name, unit, quantity, unit_price
            )
            results.append({
                "type": "work",
                "name": name,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "total": round(quantity * unit_price, 2),
            })
            continue

        # Не найдено — добавляем с нулевой ценой (пользователь заполнит)
        add_estimate_item(estimate_id, "material", name, unit, quantity, 0.0)
        results.append({
            "type": "material",
            "name": name,
            "unit": unit,
            "quantity": quantity,
            "unit_price": 0.0,
            "total": 0.0,
            "note": "⚠️ Не найдено в базе расценок",
        })

    estimate = get_estimate(estimate_id)
    return {
        "estimate_id": estimate_id,
        "project_name": project_name,
        "items": results,
        "total_materials": estimate["total_materials"],
        "total_work": estimate["total_work"],
        "total": estimate["total_overall"],
    }


def export_to_xlsx(estimate_data: dict, path: str | None = None) -> str:
    """Экспорт сметы в XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    # Заголовок
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Смета: {estimate_data['project_name']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Дата
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Создано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Заголовки таблицы
    headers = ["Тип", "Наименование", "Ед. изм.", "Кол-во", "Цена", "Сумма"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Данные
    row = 5
    for item in estimate_data["items"]:
        ws.cell(row=row, column=1, value="Материал" if item["type"] == "material" else "Работа").border = thin_border
        ws.cell(row=row, column=2, value=item["name"]).border = thin_border
        ws.cell(row=row, column=3, value=item["unit"]).border = thin_border
        ws.cell(row=row, column=4, value=item["quantity"]).border = thin_border
        ws.cell(row=row, column=5, value=item["unit_price"]).border = thin_border
        total_cell = ws.cell(row=row, column=6, value=item["total"])
        total_cell.border = thin_border
        total_cell.number_format = "#,##0.00"
        row += 1

    # Итоги
    row += 1
    bold_font = Font(bold=True, size=11)

    ws.cell(row=row, column=1, value="Итого материалы:").font = bold_font
    ws.cell(row=row, column=6, value=estimate_data["total_materials"]).font = bold_font
    ws.cell(row=row, column=6).number_format = "#,##0.00"

    row += 1
    ws.cell(row=row, column=1, value="Итого работы:").font = bold_font
    ws.cell(row=row, column=6, value=estimate_data["total_work"]).font = bold_font
    ws.cell(row=row, column=6).number_format = "#,##0.00"

    row += 1
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws.cell(row=row, column=1, value="ВСЕГО:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=6, value=estimate_data["total"]).font = Font(bold=True, size=12, color="006100")
    ws.cell(row=row, column=6).fill = total_fill
    ws.cell(row=row, column=6).number_format = "#,##0.00"

    # Ширина колонок
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

    if path is None:
        path = os.path.join(OUTPUT_DIR, f"smeta_{estimate_data['project_name']}_{datetime.now():%Y%m%d_%H%M%S}.xlsx")

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


def export_to_pdf(estimate_data: dict, path: str | None = None) -> str:
    """Экспорт сметы в PDF."""
    from weasyprint import HTML

    rows_html = ""
    for item in estimate_data["items"]:
        item_type = "Материал" if item["type"] == "material" else "Работа"
        note = f"<br><small style='color:red'>{item.get('note', '')}</small>" if item.get("note") else ""
        rows_html += f"""
        <tr>
            <td>{item_type}</td>
            <td>{item['name']}{note}</td>
            <td>{item['unit']}</td>
            <td>{item['quantity']}</td>
            <td>{item['unit_price']:,.2f}</td>
            <td>{item['total']:,.2f}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
body {{ font-family: 'DejaVu Sans', sans-serif; font-size: 11pt; margin: 2cm; }}
h1 {{ font-size: 16pt; color: #2c3e50; text-align: center; }}
.date {{ text-align: right; color: #666; font-size: 9pt; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
th {{ background: #4472C4; color: white; padding: 8px; text-align: left; }}
td {{ padding: 6px; border-bottom: 1px solid #ddd; }}
.total-row {{ font-weight: bold; }}
.grand-total {{ font-weight: bold; font-size: 12pt; color: #006100; background: #E2EFDA; }}
</style>
</head>
<body>
<h1>Смета: {estimate_data['project_name']}</h1>
<p class="date">{datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
<table>
<tr>
    <th>Тип</th>
    <th>Наименование</th>
    <th>Ед.</th>
    <th>Кол-во</th>
    <th>Цена</th>
    <th>Сумма</th>
</tr>
{rows_html}
<tr><td colspan="5" style="text-align:right"><strong>Итого материалы:</strong></td><td>{estimate_data['total_materials']:,.2f}</td></tr>
<tr><td colspan="5" style="text-align:right"><strong>Итого работы:</strong></td><td>{estimate_data['total_work']:,.2f}</td></tr>
<tr class="grand-total"><td colspan="5" style="text-align:right">ВСЕГО:</td><td>{estimate_data['total']:,.2f}</td></tr>
</table>
</body>
</html>"""

    if path is None:
        path = os.path.join(OUTPUT_DIR, f"smeta_{estimate_data['project_name']}_{datetime.now():%Y%m%d_%H%M%S}.pdf")

    os.makedirs(os.path.dirname(path), exist_ok=True)
    HTML(string=html).write_pdf(path)
    return path
